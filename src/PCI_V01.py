#!/usr/bin/env python
# coding: utf-8

# In[ ]:
import tkinter as tk
from tkinter import filedialog
import msvcrt
import os,sys
import locale
import codecs
import re
import json
import xlsxwriter
import pandas as pd
from pandas import json_normalize
import time 
from datetime import datetime
from pathlib import Path
import shutil
from shutil import rmtree
import os.path
from dateutil import parser
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
listLog=[]


#RUTA DEL PROCESOs raiz de los directorios a analizar
#locale.setlocale(locale.LC_TIME,'es_ES.UTF-8')

#path=input("1)Ingrese Ruta de busqueda: ")#RUTA DEL PROCESO raiz
#extProcess=[]
#valor=input("2)Ingrese el tipo de archivo a buscar:")
#extProcess=[valor]

#while opc=='SI':
#    valor=input("2)Ingrese Extension a procesar:")
#    extProcess.append(valor)
#    valor=input("=>Agregar otro valor (SI/NO)?: ")
#    opc=valor
    
#fileRead='.'+input("3)Ingrese el tipo de archivo a leer:")
#fileLog='.log'

#inputReadXls=r'D:\BCP\PROY_PCI\Input.xls'

#inputReadXlsx=input("4)Ingrese el archivo de entrada: ")
#output=input("5)Ingrese el archivo de salida: ")


#Funcion para capturar la EXTENSION,nombre completo, nombre sin EXTENSION y version de archivo por el nombre
def getFile(path,param):
    #param:"ext" -> extrae la extenxion del archivo
    #param:"name" -> extrae el nombre del archivo
    #param:"version" -> extrae la extenxion del archivo
    #param:"namesv" -> extrae la extenxion del archivo sin la version ejemplo (ejecutable-1.2.3.exe; result:ejecutable)
    #param:"fullname" -> extrae el nombre del archivo + la EXTENSION (ejecutable-1.2.3.exe; result:ejecutable.exe)
    path=path.split(r"/")[-1]
    if param=='ext'.lower():
        result=path.split(".")[-1]
        return result

    if param=='name'.lower():
        result=".".join(path.split(".")[0:-1])
        return result

    if param=='version'.lower():
        result=".".join(path.split("-")[-1].split(".")[0:-1])
        return result

    if param=='namesv'.lower():
        result="-".join(path.split("-")[:-1])
        return result
    
    if param=='fullname'.lower():
        result=path
        return result
    
#funcion para capturar la fecha de creación/MODIFICACION de un archivo    
def getDetailFile (path,param):
    if param=='create'.lower():#fecha creea
        result = os.path.getctime(path)
        return  datetime.fromtimestamp(result).strftime('%Y/%m/%d') 
    if param=='modify'.lower():#fecha modi
        result = os.path.getmtime(path)
        return  datetime.fromtimestamp(result).strftime('%Y/%m/%d')




  
#funcion para guardar el detalle del archivo en formato Json
def getVersionContenidoJson (path,json,read):
    #path: directorio raiz donde se encuentran los archivos
    #listArc: recibe como parametro una lista archivos (ejemplo: ["arch1.ext","arch2.ext"])
    #json: nombre del objeto
    #read: 
    #"content": lee el archivo plano y estrae los parametros a partir del contenido+\s*\d{4}
    data={}
    expfechas='\d{0,2}\/\d{0,2}\/\d{2,4}|\d{0,2}\-\d{0,2}\-\d{2,4}|{ENERO+\s*\d{4}|enero+\s*\d{4}|FEBRERO+\s*\d{4}|febrero+\s*\d{4}|MARZO+\s*\d{4}|marzo+\s*\d{4}|ABRIL+\s*\d{4}|abril+\s*\d{4}|MAYO+\s*\d{4}|mayo+\s*\d{4}|JUNIO+\s*\d{4}|junio+\s*\d{4}|JULIO+\s*\d{4}|julio+\s*\d{4}|AGOSTO+\s*\d{4}|agosto+\s*\d{4}|SETIEMBRE+\s*\d{4}|setiembre+\s*\d{4}|SEPTIEMBRE+\s*\d{4}|septiembre+\s*\d{4}|NOVIEMBRE+\s*\d{4}|noviembre+\s*\d{4}|DICIEMBRE+\s*\d{4}|diciembre+\s*\d{4}}'
    data[f'{json}']=[]
    fechadev = []
    if read=="content":
        versiones=[]
        job=''
        rowUltimaVersion=''
        #print(arc)
        text1=codecs.open(f'{path}','r',encoding='latin-1')
        for linea in text1:
            #encuentra cadena de versiones(cuerpo)
            if(re.findall('\|\|\s*[\d]\s*.*|#--\s*[\d]\s*.*|\|\s*[\d]\s*.*|--\s*\d+[\s*].*',linea)):#las lineas con las versiones
                versiones.append(linea.replace("--","").replace("#","").replace("|",""))
            #encuentra nombre del Job en el texto
            if re.findall('.(job|JOB).*:.*@\w*',linea.upper()):
                job=linea.split(":")[-1].strip()
            #Encuentra 
            if re.findall(expfechas,linea):
                fechadev.append(linea)

        try:
            count=len(versiones)-1
            while True:
                if count<0:
                    rowUltimaVersion=[''] #si no hay nada
                    break
                else:
                    if len(re.findall(expfechas,versiones[count]) )>0:
                        rowUltimaVersion=versiones[count] #si hay mas de uno
                        break
                count = count - 1    
        except:
            rowUltimaVersion=versiones #si es solo uno
            
        #print('Ultima Version:',versiones)
        #print(path)
        #print(versiones)
        #print(len(versiones)-1)
        #print( rowUltimaVersion.split()[0].strip())
        #print(fechadev)
        try:
            fechadev_list=re.findall(expfechas,rowUltimaVersion)
            if fechadev_list:
                fechadev = fechadev_list[0].strip()
            else:
                fechadev=''
        except:
            fechadev=''

        try:
            data[f'{json}'].append({
                "archivo":getFile(f'{path}','fullname'),
                "name":getFile(f'{path}','name'),
                "version": rowUltimaVersion.split()[0].strip() if len(rowUltimaVersion)>0 else '',
                "job": job,
                "tecnologia": getFile(f'{path}','ext') ,
                "fecha": getDetailFile(f'{path}','modify') ,
                "fechaDev":fechadev,
                "descripcion": re.split(expfechas,rowUltimaVersion)[-1].strip() if len(rowUltimaVersion)>0 else '' 
        })
        except:
            global listLog
            listLog.append(path)
            data[f'{json}'].append({
                "archivo":'',
                "name":'',
                "version":'',
                "job":'',
                "tecnologia":'',
                "fecha":'',
                "fechaDev":'',
                "descripcion":''
            })
        text1.close()
    return data


# Diccionario de mapeo de nombres de meses a números de mes
meses = {
    'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04', 'mayo': '05', 'junio': '06',
    'julio': '07', 'agosto': '08', 'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
}

 
def convertir_fecha(texto):
    if not isinstance(texto, str):
        texto = str(texto)

    # Expresión regular para buscar nombres de meses seguidos por un año de cuatro dígitos (ejemplo: Enero 2023)
    regex_mes_y_anio = re.compile(r'\b(\w+\w?\w?\w?\w?)\s?(\d{4})\b', re.IGNORECASE)
    # Buscar coincidencias en el texto
    match = regex_mes_y_anio.search(texto)

    if match:
        nombre_mes = match.group(1).lower()
        anio = match.group(2)
        
        # Verificar si el nombre del mes está en el diccionario
        if nombre_mes in meses:
            mes = meses[nombre_mes]
            fecha_formateada = f'{mes}/{anio}'
            return fecha_formateada

    #formatos = ['%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y', '%Y/%m/%d', '%Y-%m-%d', '%y/%m/%d', '%y-%m-%d', '%Y B%', '%B Y%']
    formatos = ['%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y','%B Y%']
    
    for formato in formatos:
        try:
            fecha = datetime.strptime(texto, formato)
            fecha_formateada = fecha.strftime('%m/%Y')
            return fecha_formateada
        except ValueError:
            pass

    return None

def resorce_path(relative_path):
    try:
        base_path=sys._MEIPASS
    except Exception:
        base_path=os.path.abspath(".")
    return os.path.join(base_path,relative_path)
def f_proceso (path,extProcess,inputReadXlsx,output):
    fileLog='.log'
    # leer Excel
    df_readExcel=pd.read_excel(inputReadXlsx,sheet_name='Sheet1',skiprows=0)
    df_readExcel_copy=df_readExcel.copy()
    #transformando
    df_readExcel_copy['EXTENSION']=df_readExcel_copy['NOMBRE'].apply(lambda x: getFile(str(x).lower().strip(),'ext') )
    df_readExcel_copy['FLAG_FILE']=df_readExcel_copy['EXTENSION'].apply(lambda x: '' if x.strip() in extProcess else '-')
    df_readExcel_copy['path']=df_readExcel_copy.apply(lambda x: path+str(x['RUTA DEL PROCESO'])+'\\'+str(x['NOMBRE']),axis=1)
    df_readExcel_copy['path_log']=df_readExcel_copy.apply(lambda x: path+str(x['RUTA DEL PROCESO'])+'\\'+getFile(str(x['NOMBRE']),'name')+fileLog,axis=1)
    df_readExcel_copy['FLAG_FILE']=df_readExcel_copy.apply(lambda x: '-' if x['FLAG_FILE']=='-' else 'Si' if (os.path.isfile(str(x['path']))) else 'No' ,axis=1 )
    df_readExcel_copy['FLAG_LOG']=df_readExcel_copy.apply(lambda x: '-' if x['FLAG_FILE']=='-' else 'Si' if (os.path.isfile(str(x['path_log']))) else 'No' ,axis=1 )
    df_readExcel_copy['NOMBRE_ENCONTRADO']=df_readExcel_copy.apply(lambda x: x['NOMBRE'] if x['FLAG_FILE']=='Si' else '' ,axis=1)
    df_readExcel_copy['RUTA ENCONTRADO']=df_readExcel_copy.apply(lambda x: x['RUTA DEL PROCESO'] if x['FLAG_FILE']=='Si' else '',axis=1 )
    df_readExcel_copy['FECHA ULTIMA MODIFICACION ENCONTRADO']=df_readExcel_copy.apply(lambda x: getVersionContenidoJson(str(x['path']),'json','content')['json'][0]['fechaDev'] if x['FLAG_FILE']=='Si' else '',axis=1 )
    df_readExcel_copy['FECHA LOG']=df_readExcel_copy.apply(lambda x: getDetailFile(str(x['path_log']),'modify') if x['FLAG_LOG']=='Si' else '',axis=1 )
    df_readExcel_copy['JOB ENCONTRADO']=df_readExcel_copy.apply(lambda x: getVersionContenidoJson(str(x['path']),'json','content')['json'][0]['job'] if x['FLAG_FILE']=='Si' else '',axis=1 )
    df_readExcel_copy['FECHANEW'] = df_readExcel_copy.apply(lambda x: convertir_fecha(x['FECHA ULTIMA MODIFICACION ENCONTRADO']) if x['FLAG_FILE'] == 'Si' else x['FECHA ULTIMA MODIFICACION ENCONTRADO'], axis=1)
    df_readExcel_copy['FECHAOLD'] = df_readExcel_copy.apply(lambda x: convertir_fecha(x['FECHA ULTIMA MODIFICACION']) if x['FLAG_FILE'] == 'Si' else x['FECHA ULTIMA MODIFICACION'], axis=1)
    df_readExcel_copy['IND_FECHA'] = df_readExcel_copy.apply(lambda x: 'Si' if x['EXTENSION']  and x['FECHANEW'] == x['FECHAOLD'] else 'NO' if x['FLAG_FILE'] == 'Si' else '', axis=1)
    df_readExcel_copy.iloc[:, [0,1,2,3,4,5,6,1,16,3,12,13,14,15,8,11,7,19]].to_excel(f"{path}\{output}.xlsx",sheet_name="Resultado",index=False)
    result="GENERADO SATISFACTORIAMENTE"
    #gnerar excel con formato de colores para las cabeceras
    orange=PatternFill(fill_type="solid",start_color="00FF6600")
    yellow=PatternFill(fill_type="solid",start_color="00FFFF00")
    green=PatternFill(fill_type="solid",start_color="0000FF00")
    wb=load_workbook(f"{path}\{output}.xlsx")
    ws=wb["Resultado"]
    max_column=ws.max_column

    for col in range(1, max_column + 1):
        if col<8:
            cell_header = ws.cell(1, col)
            cell_header.fill=orange
        elif col<14:
            cell_header = ws.cell(1, col)
            cell_header.fill=yellow
        elif col>13:
            cell_header = ws.cell(1, col)
            cell_header.fill=green  
            
    wb.save(f"{path}\{output}.xlsx")
    #print("Execute Success!!")
    #msvcrt.getch()
    #os.system("pause")
    return result
#Interfaz grafica
class Aplicacion (tk.Frame):
    def __init__(self,master=None):
        super().__init__(master)
        self.master=master
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        self.createWidgets()
    def createWidgets (self):
        self.lblWorkPath=tk.Label(text="Ruta de trabajo:")
        self.lblWorkPath.grid(row=0, column=0)
        self.txtWorkPath=tk.Entry()
        self.txtWorkPath.grid(row=0,column=1)
        def f_btnWorkPath ():
            cwd = os.getcwd()
            pathWork=filedialog.askdirectory(initialdir=cwd,title="Seleccionar directorio")
            resp=pathWork
            self.txtWorkPath.delete(0,"end")
            self.txtWorkPath.insert(0,pathWork)
            
        self.btnWorkPath=tk.Button(text="Seleccionar directorio",command=f_btnWorkPath)
        self.btnWorkPath.grid(row=0, column=2)
        
        self.chkHql=tk.IntVar()
        self.chkSql=tk.IntVar()
        self.chkPy=tk.IntVar()
        
        self.lblProcesar=tk.Label(text="Procesar:")
        self.lblProcesar.grid(row=1, column=0)
        self.chkProcesar1=tk.Checkbutton(text="hql",variable=self.chkHql)
        self.chkProcesar1.grid(row=1, column=1)
        self.lblProcesar=tk.Checkbutton(text="sql",variable=self.chkSql)
        self.lblProcesar.grid(row=1, column=2)
        self.lblProcesar=tk.Checkbutton(text="py",variable=self.chkPy)
        self.lblProcesar.grid(row=1, column=3)
        

        def f_btnFileInput ():
            cwd = os.getcwd()
            FileInput=filedialog.askopenfilename(initialdir=cwd,title="Seleccionar directorio" )
            self.txtFileInput.delete(0,"end")
            self.txtFileInput.insert(0,FileInput) 

            
        self.lblFileInput=tk.Label(text="Archivo Entrada:")
        self.lblFileInput.grid(row=2,column=0)
        self.txtFileInput=tk.Entry()
        self.txtFileInput.grid(row=2,column=1)
        self.btnFileInput=tk.Button(text="Seleccionar Archivo",command=f_btnFileInput)
        self.btnFileInput.grid(row=2, column=2)
        
        self.lblFileOutput=tk.Label(text="Nombre Archivo Salida:")
        self.lblFileOutput.grid(row=3,column=0)
        self.txtFileOutput=tk.Entry(text="Nombre Archivo Salida:")
        self.txtFileOutput.grid(row=3,column=1)
        
        def f_btnProcesar():
            listExt=[]
            if self.chkSql.get()==1:
                listExt.append("sql")
            if self.chkHql.get()==1: 
                listExt.append("hql" )
            if self.chkPy.get()==1: 
                listExt.append("py")
            #print(listExt)
            #print(self.txtWorkPath.get(),self.txtFileInput.get(),self.txtFileOutput.get() )
            f_proceso(path=self.txtWorkPath.get(),extProcess=listExt,inputReadXlsx=self.txtFileInput.get(),output=self.txtFileOutput.get())
            self.txaMsgoutput.delete('1.0',"end")
            msg="SUCCESS!!\n"
            self.txaMsgoutput.insert('1.0',msg)
            global listLog
            for val in listLog:
                self.txaMsgoutput.insert('1.0',f"{val}\n")
            self.txaMsgoutput.insert('1.0',f"Nombre Archivo: {self.txtFileOutput.get()}\n")
            self.txaMsgoutput.insert('1.0',f"Ruta Salida: {self.txtWorkPath.get()}\n")
            self.txaMsgoutput.insert('1.0',"---------------------\n")
            self.txaMsgoutput.insert('1.0',"Detalle de Ejecucion:\n")
            listLog=[]
            
        
        self.btnProcesar=tk.Button(text="Procesar",height=1,width=28,fg = 'black', bd=4,command=f_btnProcesar)
        self.btnProcesar.grid(row=4, column=0,columnspan=4)
                
        
        self.txaMsgoutput=tk.Text(height=12, width=50)
        self.txaMsgoutput.grid(row=5, column=0, columnspan=5)
        

        
        
root=tk.Tk()
root.title("Proy PCI")
root.geometry("570x350")
path_img=resorce_path(r"..\img\bcp.png")
icono=tk.PhotoImage(file=path_img)
root.iconphoto(True, icono)
app=Aplicacion(master=root)
app.mainloop()

